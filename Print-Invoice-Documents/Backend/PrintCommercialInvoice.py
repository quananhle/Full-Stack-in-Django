from django.conf import settings
import os
import psycopg2
import math
import json

from srv.database.db_actions import *
from openpyxl import *
from datetime import datetime

class PrintCommercialInvoice():
    def __init__(self, dn=None, template=None):
        self.deliveryNumber = dn
        self.fn_error = None

        self.generatedDocuments = []

        self.database = DataLayer()
        self.database_conn = self.database.connect()

    def format_template(self, template=None):
        #SHIPPER SECTION
        template.merge_cells('A7:B7')
        template.merge_cells('A8:B8')
        template.merge_cells('A9:B9')
        template.merge_cells('A10:B10')
        template.merge_cells('A11:B11')

        #SELLER SECTION
        template.merge_cells('D7:F7')
        template.merge_cells('D8:F8')
        template.merge_cells('D9:F9')
        template.merge_cells('D10:F10')
        template.merge_cells('D11:F11')

        # #SHIP TO SECTION
        template.merge_cells('A13:B13')
        template.merge_cells('A14:B14')
        
        # #SOLD TO SECTION
        template.merge_cells('D13:F13')
        template.merge_cells('D14:F14')

        #SALES INFO SECTION
        template.merge_cells('I19:J19')
        template.merge_cells('I20:J20')
        template.merge_cells('I21:J21')
        template.merge_cells('I22:J22')
        template.merge_cells('I23:J23')
        template.merge_cells('I24:J24')
        template.merge_cells('I25:J25')
        template.merge_cells('I26:J26')
        template.merge_cells('I27:J27')
        template.merge_cells('I28:J28')
        template.merge_cells('I29:J29')
        template.merge_cells('I30:J30')

    def format_dimension_row(self, template=None, row_list=[]):
        for row in row_list:
            row_str = 'A' + row + ':K' + row
            template.merge_cells(row_str)

    def write_to_template(self):
        result = {}

        try:
            docTemplatePath = settings.TEMPLATES_PATH
            genTemplatePath = settings.GEN_TEMPLATES_PATH_COMMERCIAL_INNVOICE
            get_files_path  = f"{settings.SERVER_IP}static/gen_documents/commercial_invoice/"

            templateName = os.path.join(docTemplatePath, "Commercial_Invoice_Template.xlsx")
            
            dn_hea_param = self.database.createparams(f"HEADER,{self.deliveryNumber}")
            dn_det_param = self.database.createparams(f"DETAIL,{self.deliveryNumber}")

            headerData = self.database.runfunction(self.database_conn, 'commercial_invoice_header_fn', dn_hea_param)
            headerData = self.database.runfunction(self.database_conn, 'commercial_invoice_header_fn', dn_hea_param)
            if not headerData:
                raise Exception(f"There is no data for: {self.deliveryNumber}")

            detailData = self.database.runfunction(self.database_conn, 'commercial_invoice_detail_fn', dn_det_param)

            detailLength = len(detailData)  # total amt of data to be written on invoice
            pageDetailLimit = 31.0  # current invoice limit per page (invoice has rows 20-31 available for possible data)
            pageCounter = 1
            numPages = math.ceil(detailLength / pageDetailLimit)  # determine how many pages are needed
            datacol = 'ABCDEFGHIJ'  # columns for each data entry point ABDFGHJ
            dataIterator = 0  # used for iterating through the tuple of data

            if(detailLength < pageDetailLimit):
                rowLimit = detailLength + 19  # set rowlimit as detailLength + 19 if less than 12 items since data begins on row 19
            else:
                rowLimit = 30  # set rowlimit per page as 30 based on invoice template

            currentdata = None
            get_files = []
            fileName = None

            while dataIterator < detailLength:
                dimmension_row = []
                template = load_workbook(templateName)
                template.active
                template_worksheet = template.worksheets[0]

                # Loading Header
                for _, tuple in enumerate(headerData):
                    if tuple[3] != None:
                        template_worksheet[tuple[1]] = tuple[3]
                    else:
                        template_worksheet[tuple[1]] = ''

                invoice_date= str(template_worksheet['J6'].value)
                invoice_date_list= invoice_date.split("/")
                invoice_date= str(invoice_date_list[2]+invoice_date_list[0]+invoice_date_list[1])

                new_filename = template_worksheet['J7'].value + '_' + datetime.now().strftime('%Y%m%d%H%M%S')

                for row in range(19, rowLimit):
                    currentdata = detailData[dataIterator]

                    for col, info in zip(datacol, currentdata):
                        # if info == 'DIMENSION':
                        #     dimmension_row.append(str(row))
                        #     continue
                        loc = col + str(row)
                        template_worksheet[loc] = info

                    if((dataIterator + 1) % pageDetailLimit == 0):  # saves each page as a separate file everytime the invoice is filled
                        fileName = f"CommercialInvoice_{str(new_filename)}_{str(pageCounter)}.xlsx"
                        new_filepath = os.path.join(genTemplatePath, fileName)
                        # new_filepath = os.path.join(get_files_path, fileName)
                        pageCounter = pageCounter + 1
                        self.format_template(template_worksheet)
                        self.format_dimension_row(template_worksheet,dimmension_row)
                        template.save(new_filepath)
                        get_files.append(fileName)

                    elif(dataIterator == detailLength - 1):  # saves the last page of the data
                        fileName = f"CommercialInvoice_{str(new_filename)}_{str(pageCounter)}.xlsx"
                        new_filepath = os.path.join(genTemplatePath, fileName)
                        # new_filepath = os.path.join(get_files_path, fileName)
                        self.format_template(template_worksheet)
                        self.format_dimension_row(template_worksheet,dimmension_row)
                        template.save(new_filepath)
                        get_files.append(fileName)

                    dataIterator = dataIterator + 1
                    if(dataIterator == detailLength):
                        break
            result.update({
                'fileList': get_files,
                'filesPath': get_files_path
            })
            return result

        except(Exception, psycopg2.DatabaseError, AttributeError) as error:
            result.update({
                'error': str(error).replace("'", "")
            })
            return result

        finally:
            print("PostgreSQL Connection is closed")
